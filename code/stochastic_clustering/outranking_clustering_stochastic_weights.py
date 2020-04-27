import pandas as pd
from pandas import DataFrame as df
import openpyxl as px
import numpy as np
#from plot_clusters import *
from cluster import *

from src.cluster import get_ordered_centroids_4


def init_data(excel_file):
    """
    initialises data
    :param excel_file:
    :return: actions, centroids, limites
    """
    work_book  = px.load_workbook(excel_file)
    work_sheet = work_book['data']
    df_data    = pd.DataFrame(work_sheet.values)

    # slice data to get data frames for actions, centroids, min and max
    actions    = df.to_numpy(df_data.iloc[0:189])
    centroids  = df.to_numpy(df_data.iloc[190:194])
    limites    = df.to_numpy(df_data.iloc[189:195])

    return actions, centroids, limites

def random_thresholds(excel_file):
    """
     read random thresholds (externally generated)
     :param excel_file:
     :return: p_dir,q_dir,p_inv,q_inv
     """
    work_book = px.load_workbook(excel_file)
    work_sheet_p_dir = work_book['p_dir']
    work_sheet_q_dir = work_book['q_dir']
    work_sheet_p_inv = work_book['p_inv']
    work_sheet_q_inv = work_book['q_inv']
    df_data_p_dir = pd.DataFrame(work_sheet_p_dir.values)
    df_data_q_dir = pd.DataFrame(work_sheet_q_dir.values)
    df_data_p_inv = pd.DataFrame(work_sheet_p_inv.values)
    df_data_q_inv = pd.DataFrame(work_sheet_q_inv.values)

    # slice data to get data frames for thresholds
    pdir = df.to_numpy(df_data_p_dir.iloc[0:3000])
    qdir = df.to_numpy(df_data_q_dir.iloc[0:3000])
    pinv = df.to_numpy(df_data_p_inv.iloc[0:3000])
    qinv = df.to_numpy(df_data_q_inv.iloc[0:3000])

    return pdir,qdir,pinv,qinv




def random_weights(excel_file):
    """
     read random weights (externally generated)
     :param excel_file:
     :return: w
     """
    work_book = px.load_workbook(excel_file)
    work_sheet_w = work_book['weights']
    df_data_w = pd.DataFrame(work_sheet_w.values)

    # slice data to get data frames for thresholds
    w = df.to_numpy(df_data_w.iloc[0:1000])

    return w

def get_weights():
    """
    get criteria weights
    :return: w
    """
    w = [0.333, 0.333, 0.334]  # pesos de los criterios
    return w

def get_umbrales():
    """
    umbrales de preferencia directos e inversos de cada criterio
    :return: p_dir, q_dir, p_inv, q_inv
    """


    p_dir = [0.19,0.14,0.10]
    q_dir = [0.09,0.07,0.05]
    p_inv = [0.19,0.14,0.10]
    q_inv = [0.09,0.07,0.05]

    return p_dir, q_dir, p_inv, q_inv


def get_metrics(actions, limites):
    """
    Get lengths of input data
    :param actions:
    :param limites:
    :return: n_acc, n_cri, n_limites
    """
    n_acc = np.size(actions, 0)  # number of acciones
    n_cri = np.size(actions, 1)  # number if criteria
    n_lim = np.size(limites, 0)  # number of limits
    return n_acc, n_cri, n_lim

def conc_p_directa_actions(actions, p_dir, q_dir):
    n_acc = np.size(actions, 0)  # number of acciones
    n_cri = np.size(actions, 1)  # number if criteria
    cpda  = np.zeros((n_acc, n_acc, n_cri))

    # calcula indice de concordancia parcial directo
    for h in range(0, n_cri):
        # mueve j en las filas del arreglo de perfiles de categorias
        for j in range(0, n_acc):
            # mueve i en las filas del arreglo de acciones
            for i in range(0, n_acc):
                if actions[i][h] - actions[j][h] > p_dir[h]:
                    cpda[j][i][h] = 0
                else:
                    if actions[i][h] - actions[j][h] <= q_dir[h]:
                        cpda[j][i][h] = 1
                    else:
                        cpda[j][i][h] = 1.0 * (actions[j][h] - actions[i][h] + p_dir[h]) / (p_dir[h] - q_dir[h])
    return cpda


def conc_p_inversa_actions(actions, p_inv, q_inv):
    """
    calcula la concordancia parcial inversa
    :param actions:
    :param p_inv:
    :param q_inv:
    :return:
    """
    n_acc = np.size(actions, 0)  # number of acciones
    n_cri = np.size(actions, 1)  # number if criteria

    cpia  = np.zeros((n_acc, n_acc, n_cri))  # cpi array to store values

    for h in range(0, n_cri):
        # calcula indice de concordancia parcial inverso
        #mueve i en las filas del arreglo de acciones
        for i in range(0, n_acc):
            #mueve j en las filas del arreglo de perfiles de categorias
            for j in range(0, n_acc):
                if actions[j][h] - actions[i][h] > p_inv[h]:
                    cpia[i][j][h] = 0
                else:
                    if actions[j][h] - actions[i][h] <= q_inv[h]:
                        cpia[i][j][h] = 1
                    else:
                        cpia[i][j][h] = 1.0 * (actions[i][h] - actions[j][h] + p_inv[h]) / (p_inv[h] - q_inv[h])
    return cpia


def conc_p_directa(actions, limites, p_dir, q_dir):
    """
    calcula la concordancia parcial directa
    :param actions:
    :param limites:
    :param p_dir:
    :param q_dir:
    :return: cod
    """
    n_acc = np.size(actions, 0)  # number of acciones
    n_cri = np.size(actions, 1)  # number if criteria
    n_lim = np.size(limites, 0)  # number of limits
    cpd = np.zeros((n_lim, n_acc, n_cri))

    # calcula indice de concordancia parcial directo
    for h in range(0, n_cri):
        # mueve j en las filas del arreglo de perfiles de categorias
        for j in range(0, n_lim):
            # mueve i en las filas del arreglo de acciones
            for i in range(0, n_acc):
                if actions[i][h] - limites[j][h] > p_dir[h]:
                    cpd[j][i][h] = 0
                else:
                    if actions[i][h] - limites[j][h] <= q_dir[h]:
                        cpd[j][i][h] = 1
                    else:
                        cpd[j][i][h] = 1.0 * (limites[j][h] - actions[i][h] + p_dir[h]) / (p_dir[h] - q_dir[h])
    return cpd

def conc_p_inversa(actions, limites, p_inv, q_inv):
    """
    calcula la concordancia parcial inversa
    :param actions:
    :param limites:
    :param p_inv:
    :param q_inv:
    :return:
    """
    n_acc = np.size(actions, 0)  # number of acciones
    n_cri = np.size(actions, 1)  # number if criteria
    n_lim = np.size(limites, 0)  # number of limits

    cpi   = np.zeros((n_acc, n_lim, n_cri))  # cpi array to store values

    for h in range(0, n_cri):
        # calcula indice de concordancia parcial inverso
        #mueve i en las filas del arreglo de acciones
        for i in range(0, n_acc):
            #mueve j en las filas del arreglo de perfiles de categorias
            for j in range(0, n_lim):
                if limites[j][h] - actions[i][h] > p_inv[h]:
                    cpi[i][j][h] = 0
                else:
                    if limites[j][h] - actions[i][h] <= q_inv[h]:
                        cpi[i][j][h] = 1
                    else:
                        cpi[i][j][h] = 1.0 * (actions[i][h] - limites[j][h] + p_inv[h]) / (p_inv[h] - q_inv[h])
    return cpi

def concordancia_I_actions(cpia, n_acc, n_cri, w):
    """
    calcula la concordancia inversa
    :param cpi:
    :param n_acc:
    :param n_cri:
    :param w:
    :return: sigma_I
    """
    sigma_I = np.zeros((n_acc, n_acc))
    for i in range(0, n_acc):
        for j in range(0, n_acc):
            x = 0.0
            for h in range(0, n_cri):
                x = x + w[h] * cpia[i][j][h]
            sigma_I[i][j] = x
    return sigma_I


def concordancia_D_actions(cpda, n_acc,  n_cri, w):
    """
    calcula la concordancia directa
    :param cpd:
    :param n_acc:
    :param n_cri:
    :param w:
    :return: sigma_D
    """
    sigma_D = np.zeros((n_acc, n_acc))
    for i in range(0, n_acc):
        for j in range(0, n_acc):
            x = 0.0
            for h in range(0, n_cri ):
                x = x + w[h] * cpda[i][j][h]
            sigma_D[i][j] = x
    return sigma_D


def concordancia_I(cpi, n_acc, n_lim, n_cri, w):
    """
    calcula la concordancia inversa
    :param cpi:
    :param n_acc:
    :param n_lim:
    :param n_cri:
    :param w:
    :return: sigma_I
    """
    sigma_I = np.zeros((n_acc, n_lim))
    for i in range(0, n_acc):
        for j in range(0, n_lim):
            x = 0.0
            for h in range(0, n_cri):
                x = x + w[h] * cpi[i][j][h]
            sigma_I[i][j] = x
    return sigma_I


def concordancia_D(cpd, n_acc, n_lim, n_cri, w):
    """
    calcula la concordancia directa
    :param cpd:
    :param n_acc:
    :param n_lim:
    :param n_cri:
    :param w:
    :return: sigma_D
    """
    sigma_D = np.zeros((n_lim, n_acc))
    for i in range(0, n_lim):
        for j in range(0, n_acc):
            x = 0.0
            for h in range(0, n_cri ):
                x = x + w[h] * cpd[i][j][h]
            sigma_D[i][j] = x
    return sigma_D


def regla_desc(categoria, belonging, n_acc, n_lim, sigma_I, sigma_D, lam):
    """
    implemena la regla descendente
    :param n_acc:  number of acciones
    :param n_lim: number of limits
    :param sigma_I:
    :param sigma_D:
    :param lam:
    :return: categoria
    """
    for i in range(0, n_acc):
        j = n_lim - 1
        while True:
            if j == 0 :
                categoria[i][1] = 1
                belonging[i]=1
                break
            else:
                if sigma_I[i][j] >= lam:
                    if j == n_lim-1 :
                        categoria[i][n_lim - 2]=1
                        belonging[i]=n_lim - 2
                    else:
                        if min(sigma_I[i][j], sigma_D[j][i]) > min(sigma_I[i][j + 1], sigma_D[j + 1][i]):
                            categoria[i][j] = 1
                            belonging[i]=j
                        else:
                            if j == (n_lim - 2):
                                categoria[i][n_lim - 2] = 1
                                belonging[i]=n_lim - 2
                            else:
                                categoria[i][j + 1] = 1
                                belonging[i]=j+1
                    break
            j = j-1

    return categoria

def sumAscendente(freq_acceptability,categoria,n_lim,n_acc):
    for i in range(0,n_acc):
        for j in range(0,n_lim):
            freq_acceptability[i][j]=freq_acceptability[i][j]+categoria[i][j]
    return freq_acceptability

def aceptabilidadDescendente(iter_stochastic,freq_acceptability,n_lim,n_acc):
    for i in range(0,n_acc):
        for j in range(0,n_lim):
            freq_acceptability[i][j]=freq_acceptability[i][j]/float(iter_stochastic)
            print(str("%.2f" %  round(freq_acceptability[i][j],2)),"\t", end="")
        print (" ")
    print ("")

def check_separability(limites,n_lim,n_cri):
    no_check='False'
    for i in range(2,n_lim):
        for j in range(0,n_cri):
            if limites[i][j]<limites[i-1][j]:
                no_check='True'
    return no_check



def perform_outranking(actions, limites, lam, beta, iter,p_dir, q_dir, p_inv, q_inv,w,iter_stochastic):
    #w = get_weights()
    freq_no_check=0.0
    #p_dir, q_dir, p_inv, q_inv = get_umbrales()
    n_acc = np.size(actions, 0)  # number of acciones
    n_cri = np.size(actions, 1)  # number of criteria
    n_lim = np.size(limites, 0)  # number of limits
    freq_acceptability = np.zeros((n_acc, n_lim))
    # -------------------------------------------------------
    # Calcula concordancia global directa e inversa entre cada par de acciones
    # cpda = conc_p_directa_actions(actions, p_dir[iter_stochastic], q_dir[iter_stochastic])
    # cpia = conc_p_inversa_actions(actions, p_inv[iter_stochastic], q_inv[iter_stochastic])
    # sigma_D_a = concordancia_D_actions(cpda, n_acc, n_cri, w)
    # sigma_I_a = concordancia_I_actions(cpia, n_acc, n_cri, w)

    for l in range (0,iter_stochastic):
        print (l)
        for k in range(0, iter):
            # calcula concordancia parcial directa e inversa (formulas (1) y (2)
            cpd = conc_p_directa(actions, limites, p_dir, q_dir)
            cpi = conc_p_inversa(actions, limites, p_inv, q_inv)

            # calcula concordancia global directa e inversa (formulas (3) y (4)
            sigma_D = concordancia_D(cpd, n_acc, n_lim, n_cri, w[l])
            sigma_I = concordancia_I(cpi, n_acc, n_lim, n_cri, w[l])


            # determina categoria de cada accion, usando regla descendente
            categoria = np.zeros((n_acc, n_lim), dtype=int)
            belonging = np.zeros((n_acc), dtype=int)

            categoria = regla_desc(categoria,belonging, n_acc, n_lim, sigma_I, sigma_D, lam)

            # actualiza centroides por el metodo de los promedios
            #limites = get_ordered_centroids(categoria, actions, limites, p_dir,p_inv,n_acc, n_lim, n_cri)

            # actualiza centroides por el metodo de los promedios limitados a acciones dentro de trapezoides
            #limites = get_ordered_centroids_2(categoria, actions, limites, p_dir,p_inv,n_acc, n_lim, n_cri)

            #determina los nuevos centroides de categoria, usando técnica de Fernandez et al. (2010)
            #limites=get_ordered_centroids_3(categoria,n_lim,n_acc,sigma_D_a,sigma_I_a,beta,n_cri,limites,actions)

            #determina los nuevos centroides de categoria, usando técnica cuchufletina

            #limitesold=limites
            limites=get_ordered_centroids_4(belonging,limites,n_lim,n_cri,p_dir,p_inv,actions)


        freq_acceptability=sumAscendente(freq_acceptability, categoria, n_lim, n_acc)

        #counts how many times the weak separability condition is not satisfied
        #print(limites)
        if check_separability(limites, n_lim, n_cri) == 'True':
            freq_no_check = freq_no_check + 1

        # print ("aqui")
        # print (freq_acceptability)

            # print('--------------- ITERACION: {} -------------'.format(k+1))
            # print('<CATEGORIAS>')
            #print(categoria)
            # for j in range (1,n_lim-1):
            #     for i in range (0,n_acc):
            #         if categoria[i][j]==1:
            #             print (j,"-", i,":",actions[i][0],actions[i][1],actions[i][2],actions[i][3],actions[i][4])

            # CALL PLOTTING HERE
            # print('<CENTROIDES>')
            # np.set_printoptions(precision=2)
            # print(limites[:])
            # for i in range(0,n_acc):
            #     print (categoria[i][0],categoria[i][1],categoria[i][2],categoria[i][3],categoria[i][4],categoria[i][5])
    #############################################
    # plot_centroids(actions, limites, k) # experiments with plotting centroids
    # print('<ACTION>')
    # print(actions[:,0])
    # print(limites[:,2:3])

    aceptabilidadDescendente(iter_stochastic,freq_acceptability,n_lim,n_acc)
    print (freq_no_check/iter_stochastic)
    return 0


#########  MAIN ###############
def main():
    actions, centroids, limites = init_data('HDI.xlsx')
    #p_dir, q_dir, p_inv, q_inv=random_thresholds('random_umbrales_2.xlsx')
    w=random_weights('Weights.xlsx')
    lam  = 0.8
    beta=0.4
    iter_stochastic=1000
    iter = 30

    p_dir, q_dir, p_inv, q_inv=get_umbrales()
    perform_outranking(actions, limites,lam,beta, iter, p_dir, q_dir, p_inv, q_inv,w,iter_stochastic)

    print (centroids)
if __name__ == '__main__':
    main()




