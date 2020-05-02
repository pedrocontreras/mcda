import numpy as np
import pandas as pd
from pandas import DataFrame as df
from pathlib import Path
import openpyxl as px

def folder(excel_folder):
    """
    initialises data
    :param excel_folder:
    :return: data_folder
    """
    data_folder = Path(excel_folder)
    return data_folder

def init_data(excel_file,i_actions,f_actions,i_centroids,f_centroids,i_limites,f_limites):
    """
    initialises data
    :param excel_file:
    :return: actions, centroids, limites
    """
    work_book  = px.load_workbook(excel_file)
    work_sheet = work_book['data']
    df_data    = pd.DataFrame(work_sheet.values)

    # slice data to get data frames for actions, centroids, min and max
    # Acciones de HDI
    # actions    = df.to_numpy(df_data.iloc[0:189])
    # centroids  = df.to_numpy(df_data.iloc[190:194])
    # limites    = df.to_numpy(df_data.iloc[189:195])

    #Acciones de SSI
    actions    = df.to_numpy(df_data.iloc[i_actions:f_actions])
    centroids  = df.to_numpy(df_data.iloc[i_centroids:f_centroids])
    limites    = df.to_numpy(df_data.iloc[i_limites:f_limites])

    return actions, centroids, limites

def get_metrics(actions, limites):
    """
    Get lengths of input data
    :param actions:
    :param limites:
    :return: n_acc, n_cri, n_limites
    """
    n_acc = np.size(actions, 0)  # number of acciones
    n_cri = np.size(actions, 1)  # number if criteria
    n_lim = np.size(limites, 0)  # number of limits
    return n_acc, n_cri, n_lim


def random_thresholds(excel_file,i_random,f_random):
    """
     read random thresholds (externally generated)
     :param excel_file:
     :return: p_dir,q_dir,p_inv,q_inv
     """
    work_book = px.load_workbook(excel_file)
    work_sheet_p_dir = work_book['p_dir']
    work_sheet_q_dir = work_book['q_dir']
    work_sheet_p_inv = work_book['p_inv']
    work_sheet_q_inv = work_book['q_inv']
    df_data_p_dir = pd.DataFrame(work_sheet_p_dir.values)
    df_data_q_dir = pd.DataFrame(work_sheet_q_dir.values)
    df_data_p_inv = pd.DataFrame(work_sheet_p_inv.values)
    df_data_q_inv = pd.DataFrame(work_sheet_q_inv.values)

    # slice data to get data frames for thresholds
    pdir = df.to_numpy(df_data_p_dir.iloc[i_random:f_random])
    qdir = df.to_numpy(df_data_q_dir.iloc[i_random:f_random])
    pinv = df.to_numpy(df_data_p_inv.iloc[i_random:f_random])
    qinv = df.to_numpy(df_data_q_inv.iloc[i_random:f_random])

    return pdir,qdir,pinv,qinv

def random_weights(excel_file,i_wrandom,f_wrandom):
    """
     read random weights (externally generated)
     :param excel_file:
     :return: w
     """
    work_book = px.load_workbook(excel_file)
    work_sheet_w = work_book['weights']
    df_data_w = pd.DataFrame(work_sheet_w.values)

    # slice data to get data frames for thresholds
    w = df.to_numpy(df_data_w.iloc[i_wrandom:f_wrandom])

    return w
